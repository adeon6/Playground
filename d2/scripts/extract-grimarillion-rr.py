from __future__ import annotations

import json
import os
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_NAME = "Build Calculator for Resist Reduction.xlsx"
SHEET_NAME = "Grimarillion Calculator"
OUTPUT_FILE = Path(__file__).resolve().parents[1] / "data" / "grimarillion-rr.js"
MASTERY_ROWS = range(5, 37)


def find_workbook(root: Path) -> Path:
    for directory, _, filenames in os.walk(root):
        if WORKBOOK_NAME in filenames:
            return Path(directory) / WORKBOOK_NAME
    raise FileNotFoundError(f"Could not find {WORKBOOK_NAME!r} under {root}")


def normalize_header(column: int, header: str | None) -> str | None:
    if header is None:
        return None
    if header == "Elemental(A":
        return "Elemental(A)"
    if header == "Other (B)":
        return "Other(B)"
    if 28 <= column <= 38:
        return f"{header[:-1]}(C)"
    return header


def load_masteries(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook[SHEET_NAME]
    headers = [normalize_header(column, sheet.cell(row=3, column=column).value) for column in range(1, 40)]
    masteries = []

    for row in MASTERY_ROWS:
        mastery = {"name": sheet.cell(row=row, column=2).value, "notes": [], "A": {}, "B": {}, "C": {}}
        for column in range(3, 40):
            header = headers[column - 1]
            cell = sheet.cell(row=row, column=column)
            if cell.value is None:
                continue
            if header == "Other(B)":
                mastery["notes"].append(str(cell.value))
                continue

            numeric = float(cell.value)
            value = round(numeric * 100, 2) if cell.number_format == "0%" else round(numeric, 2)
            damage_type, category = header.split("(")
            mastery[category.rstrip(")")][damage_type] = value

        masteries.append(mastery)

    return masteries


def main() -> None:
    workbook_path = find_workbook(Path.home() / "OneDrive" / "Documents")
    dataset = {
        "sourceWorkbook": WORKBOOK_NAME,
        "sheet": SHEET_NAME,
        "damageTypes": [
            "Fire",
            "Lightning",
            "Cold",
            "Elemental",
            "Aether",
            "Chaos",
            "Vitality",
            "Poison",
            "Physical",
            "Pierce",
            "Bleeding",
        ],
        "masteries": load_masteries(workbook_path),
        "rules": {
            "A": "Pick the highest applicable source.",
            "B": "Add all applicable sources together.",
            "C": "Pick the highest applicable source.",
        },
    }

    OUTPUT_FILE.write_text(
        "window.grimarillionData = " + json.dumps(dataset, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
